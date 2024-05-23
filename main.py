from fastapi import FastAPI, Request, HTTPException, WebSocket
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.security import HTTPBasicCredentials, HTTPBasic
from fastapi.staticfiles import StaticFiles
import secrets
from pydantic import BaseModel, ValidationError
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from typing import Optional
import websockets.exceptions
import numpy as np
import os
import uvicorn

security = HTTPBasic()

app = FastAPI()

# Mount the directory containing static files
if not os.path.exists("static"):
    os.makedirs("static")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Define expected column names
expected_user_columns = ["username", "password", "email", "age", "gender", "mobile_number"]
expected_message_columns = ["sender", "receiver", "message", "timestamp"]

# Load existing data from Excel files
users_wb = load_workbook("users.xlsx")
users_sheet = users_wb.active
users_data = users_sheet.values
users_header = next(users_data)  # Get the header row
users_df = pd.DataFrame(users_data, columns=users_header)  # Create DataFrame with header row

# If the number of columns doesn't match the expected number,
# assume that the first row contains data and add default column names
if len(users_df.columns) != len(expected_user_columns):
    missing_columns = len(expected_user_columns) - len(users_df.columns)
    users_df = pd.concat([users_df, pd.DataFrame(columns=[f"missing_column_{i}" for i in range(missing_columns)])], axis=1)

# Ensure only the expected columns are retained
users_df = users_df[expected_user_columns]

# Similarly for messages_df
messages_wb = load_workbook("messages.xlsx")
messages_sheet = messages_wb.active
messages_data = messages_sheet.values
messages_header = next(messages_data)  # Get the header row
messages_df = pd.DataFrame(messages_data, columns=messages_header)  # Create DataFrame with header row

# If the number of columns doesn't match the expected number,
# assume that the first row contains data and add default column names
if len(messages_df.columns) != len(expected_message_columns):
    missing_columns = len(expected_message_columns) - len(messages_df.columns)
    messages_df = pd.concat([messages_df, pd.DataFrame(columns=[f"missing_column_{i}" for i in range(missing_columns)])], axis=1)

# A dictionary to store logged-in users and their tokens
active_users = {}

class UserRegistration(BaseModel):
    username: str
    password: str
    email: Optional[str] = None
    age: Optional[int] = None
    gender: Optional[str] = None
    mobile_number: Optional[str] = None

class Message(BaseModel):
    sender: str
    receiver: str
    message: str

class LoginRequest(BaseModel):
    username: str
    password: str

class ConnectionManager:
    def __init__(self):
        self.active_connections = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            await connection.send_text(message)

manager = ConnectionManager()

@app.post("/login/")
def login_user(login_request: LoginRequest):
    user_data = users_df[users_df["username"] == login_request.username]
    if user_data.empty or user_data["password"].iloc[0] != login_request.password:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = secrets.token_urlsafe(16)  # Generate a random token
    active_users[token] = login_request.username  # Store the token with the username
    response = RedirectResponse(url="/landing-page", status_code=302)
    response.set_cookie(key="token", value=token)  # Set the token in a cookie
    return response

@app.middleware("http")
async def authenticate_user(request: Request, call_next):
    if request.url.path == "/landing-page":
        token = request.cookies.get("token")
        if not token or token not in active_users:
            # User is not authenticated, redirect to the login page
            return RedirectResponse(url="/", status_code=302)
    response = await call_next(request)
    return response

@app.get("/", response_class=HTMLResponse)
async def login_page():
    return open("index.html").read()

@app.get("/landing-page", response_class=HTMLResponse)
async def landing_page(request: Request):
    token = request.cookies.get("token")
    if token in active_users:
        # User is authenticated, allow access to the landing page
        return open("landing-page.html").read()
    else:
        # User is not authenticated, redirect to the login page
        return RedirectResponse(url="/", status_code=302, headers={"Location": "/"})

@app.get("/register/", response_class=HTMLResponse)  # Allow GET method for register page
async def register_page():
    return open("register.html").read()

@app.post("/register/")
async def register_user(user: UserRegistration):
    try:
        print("Received payload:", user.dict())
        global users_df  # Declare as global
        if user.username in users_df["username"].values:
            raise HTTPException(status_code=400, detail="Username already taken")

        # Store the password as plain text
        plain_password = user.password

        # Create a new user entry, using default values for optional fields if they are None
        new_user = pd.DataFrame({
            "username": [user.username],
            "password": [plain_password],
            "email": [user.email or ""],
            "age": [user.age or None],
            "gender": [user.gender or ""],
            "mobile_number": [user.mobile_number or ""]
        })

        # Append new user to the users DataFrame and save to Excel
        users_df = pd.concat([users_df, new_user], ignore_index=True)
        users_df.to_excel("users.xlsx", index=False)  # Save updated data to Excel file

        return {"username": user.username}
    except ValidationError as e:
        print(f"Validation error: {e.json()}")
        raise HTTPException(status_code=422, detail=e.errors())
    except HTTPException as http_exc:
        print(f"HTTP error: {http_exc.detail}")
        raise http_exc
    except Exception as e:
        print(f"Unexpected error: {e}")
        raise HTTPException(status_code=500, detail="Internal Server Error")

class UserProfile(BaseModel):
    username: str
    email: str
    age: int
    gender: str
    mobile_number: str

@app.get("/profile/{username}/", response_model=UserProfile)
async def get_user_profile(username: str, request: Request):  # Added request: Request parameter
    user_profile = users_df[users_df["username"] == username]
    if user_profile.empty:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_data = user_profile.to_dict(orient="records")[0]
    return user_data

# Function to replace NaN values with a placeholder
def replace_nan_with_placeholder(data):
    if isinstance(data, dict):
        return {k: replace_nan_with_placeholder(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [replace_nan_with_placeholder(item) for item in data]
    elif isinstance(data, float) and np.isnan(data):
        return None  # Replace NaN with None
    else:
        return data

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Receive the message from the client
            data = await websocket.receive_text()
            
            # Broadcast the received message to all connected clients
            await manager.broadcast(data)
    except websockets.exceptions.ConnectionClosedError:
        # Disconnect the WebSocket when the connection is closed
        manager.disconnect(websocket)

@app.post("/send_message/")
def send_message(message: Message):
    global messages_df  # Declare as global
    if message.sender in users_df["username"].values and message.receiver in users_df["username"].values:
        new_message = pd.DataFrame({
            "sender": [message.sender],
            "receiver": [message.receiver],
            "message": [message.message],
            "timestamp": [datetime.now()]
        })
        messages_df = pd.concat([messages_df, new_message], ignore_index=True)
        messages_df.to_excel("messages.xlsx", index=False)  # Save updated data to Excel file
        return {"status": "Message sent"}
    else:
        raise HTTPException(status_code=404, detail="User not found")

@app.get("/messages/{username}/")
def get_messages(username: str):
    user_messages = messages_df[(messages_df['sender'] == username) | (messages_df['receiver'] == username)]
    if user_messages.empty:
        raise HTTPException(status_code=404, detail="No Messages to display")
    return user_messages.to_dict(orient="records")

@app.get("/messages/{sender}/{receiver}/")
def get_messages_between(sender: str, receiver: str):
    user_messages = messages_df[(messages_df['sender'] == sender) & (messages_df['receiver'] == receiver) | 
                                (messages_df['sender'] == receiver) & (messages_df['receiver'] == sender)]
    if user_messages.empty:
        raise HTTPException(status_code=404, detail="No messages found between the specified sender and receiver")
    return user_messages.to_dict(orient="records")

@app.get("/api/users/")
async def get_users():
    # Ignore the first row and empty rows
    users_df_filtered = users_df.iloc[0:]
    users = users_df_filtered["username"].tolist()
    return JSONResponse(content={"users": users})

@app.get("/api/logged_in_user/")
async def get_logged_in_user(request: Request):
    token = request.cookies.get("token")
    if token in active_users:
        return JSONResponse(content={"username": active_users[token]})
    else:
        raise HTTPException(status_code=401, detail="User not authenticated")

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
